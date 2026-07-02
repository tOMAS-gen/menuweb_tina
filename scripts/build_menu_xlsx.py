#!/usr/bin/env python3
"""Genera menu.xlsx y assets/menu-fallback.json desde los datos del menú.
Editar este archivo y volver a correrlo es la forma de regenerar ambos
artefactos desde una única fuente si hace falta reconstruirlos."""

import json
import os

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

HEADERS_ROW = ["nombre", "descripcion", "disponible", "precio", "precio_2", "imagen"]
HEADERS_CAFE = ["nombre", "disponible", "precio_s", "precio_m", "precio_l", "precio_xl"]

CAFE_TRADICIONAL = [
    {"nombre": "Expresso", "precio_s": 3300, "precio_m": 4200, "precio_l": 4800, "precio_xl": 5500},
    {"nombre": "Doble Expresso", "precio_s": 3300, "precio_m": 4200, "precio_l": 4800, "precio_xl": 5500},
    {"nombre": "Café con leche", "precio_s": 3300, "precio_m": 4200, "precio_l": 4800, "precio_xl": 5500},
    {"nombre": "Cortado", "precio_s": 3300, "precio_m": 4200, "precio_l": 4800, "precio_xl": 5500},
    {"nombre": "Lágrima", "precio_s": 3300, "precio_m": 4200, "precio_l": 4800, "precio_xl": 5500},
    {"nombre": "Expresso Panna", "precio_s": 3300, "precio_m": 4200, "precio_l": 4800, "precio_xl": 5500},
    {"nombre": "Ice Coffee", "precio_s": 3300, "precio_m": 4200, "precio_l": 4800, "precio_xl": 5500},
    {"nombre": "Expresso Tonic", "precio_s": 3300, "precio_m": 4200, "precio_l": 4800, "precio_xl": 5500},
]

SECTORES = {
    "Café de Especialidad": [
        {"nombre": "Capuccino", "descripcion": "Clásico / Frambuesa", "precio": 6400},
        {"nombre": "Café Moccha Vienés", "precio": 6400},
        {"nombre": "Caramel Macchiato", "precio": 6400},
        {"nombre": "Macchiato Avellanas", "precio": 6400},
        {"nombre": "Café Irlandés", "precio": 6400},
        {"nombre": "Latte Vainilla", "precio": 6400},
        {"nombre": "Affogato", "precio": 6400},
        {"nombre": "Iced Coffee", "descripcion": "Pistacho / Coco", "precio": 6400},
        {"nombre": "Café Tina", "precio": 6400},
        {"nombre": "Café Bombón", "precio": 6400},
        {"nombre": "Café caribeño", "precio": 6400},
        {"nombre": "Submarino", "precio": 6400},
        {"nombre": "Bayleis", "precio": 6400},
    ],
    "Materos": [
        {"nombre": "Jaguar / Pampa / Paradise", "descripcion": "mate + blend yerba a elección + termo 80°", "precio": 4600},
        {"nombre": "Opción Tereré", "descripcion": "mate + blend yerba a elección + termo 80°", "precio": 4900},
    ],
    "Frozzen": [
        {"nombre": "Banana con dulce de leche", "precio": 6600},
        {"nombre": "Arándanos", "precio": 6600},
        {"nombre": "Frutos rojos", "precio": 6600},
        {"nombre": "Frutillas con crema", "precio": 6600},
        {"nombre": "Mousse de limón", "precio": 6600},
    ],
    "Té en Hebras": [
        {"nombre": "Tetera para compartir", "precio": 6000},
        {"nombre": "Tetera individual", "precio": 3300},
        {"nombre": "Té / Yerbeado", "precio": 3300},
        {"nombre": "Matcha latte", "descripcion": "té verde en polvo con leche", "precio": 6400},
    ],
    "Frapucchino": [
        {"nombre": "Café bombón", "precio": 6600},
        {"nombre": "Mocha Vainilla Latte", "precio": 6600},
        {"nombre": "Caramel Avellanas", "precio": 6600},
        {"nombre": "Frambuesa", "precio": 6600},
    ],
    "Bebidas": [
        {"nombre": "Agua / Agua Saborizada", "precio": 4100},
        {"nombre": "Gaseosas", "descripcion": "Línea Pepsi", "precio": 4100},
        {"nombre": "Jugo natural de naranja", "precio": 6000, "precio_2": 7600},
        {"nombre": "Pomelada", "precio": 8700},
        {"nombre": "Limonada", "descripcion": "Clásica / Menta jengibre / Arándanos / Maracuyá / Cereza", "precio": 8700, "precio_2": 9600},
    ],
    "Pastelería Propia": [
        {"nombre": "Brownie con frutos rojos", "precio": 7100},
        {"nombre": "Crumble", "precio": 7100},
        {"nombre": "Cheesecake", "precio": 7100},
        {"nombre": "Balcarce", "precio": 7100},
        {"nombre": "Selva Negra", "precio": 7100},
        {"nombre": "Lemon Pie", "precio": 7100},
        {"nombre": "Limón Cocado", "precio": 7100},
        {"nombre": "Tiramisú", "precio": 7100},
    ],
    "Antojos": [
        {"nombre": "Alfajores Premium", "descripcion": "Pistacho / Red Velvet / Chocotorta", "precio": 6600},
        {"nombre": "Alfajor Maicena / Miel", "precio": 4200},
        {"nombre": "Financieros", "descripcion": "4 unidades", "precio": 2800},
        {"nombre": "Madeleine", "descripcion": "3 unidades", "precio": 2800},
        {"nombre": "Cookies", "precio": 2800},
        {"nombre": "Macarons", "precio": 7000},
        {"nombre": "Budín", "precio": 3500},
        {"nombre": "Scones", "descripcion": "3 unidades", "precio": 3600},
        {"nombre": "Tortita", "precio": 850},
        {"nombre": "Tostado jamón y queso", "precio": 9200},
        {"nombre": "Medialuna", "descripcion": "opc. jamón y queso", "precio": 1400, "precio_2": 3300},
        {"nombre": "Tostón con omelette", "precio": 9200},
        {"nombre": "Bruschetta Tina", "precio": 9200},
        {"nombre": "Galletas Bretón", "descripcion": "3 unidades", "precio": 2100},
    ],
    "Focaccias Rellenas": [
        {"nombre": "Serrano", "descripcion": "rúcula, jamón crudo y nueces", "precio": 10800},
        {"nombre": "Azul", "descripcion": "queso untable, queso roquefort y tomate hidratado", "precio": 10800},
        {"nombre": "Gourmet", "descripcion": "peras, queso azul, rúcula, nueces y miel", "precio": 10800},
        {"nombre": "Capresse", "descripcion": "tomate, queso y albahaca", "precio": 10800},
        {"nombre": "Tradicional", "descripcion": "jamón, queso, lechuga y tomate", "precio": 10800},
        {"nombre": "Fugazza", "descripcion": "queso y cebolla caramelizada", "precio": 10800},
    ],
    "Pastelería Moderna": [
        {"nombre": "Hawaiana", "descripcion": "mousse de coco y corazón de ananá", "precio": 7500, "imagen": "torta-hawaiana.png"},
        {"nombre": "Agus Tina", "descripcion": "mousse de chocolate, centro de naranja, terciopelo de chocolate", "precio": 7500, "imagen": "torta-agus-tina.png"},
        {"nombre": "Valen Tina", "descripcion": "profiterol relleno de cheesecake de pistacho, centro de praline de pistacho", "precio": 7500, "imagen": "torta-valen-tina.png"},
        {"nombre": "Capricho", "descripcion": "cremoso de chocolate, centro de maracuyá y frambuesa, terciopelo de chocolate blanco", "precio": 7500, "imagen": "torta-capricho.png"},
        {"nombre": "Brownie boom", "descripcion": "húmedo de chocolate, pasta de maní y dulce de leche, bañado en chocolate y crocante de maní", "precio": 7500, "imagen": "torta-brownie-boom.png"},
        {"nombre": "Banofee", "descripcion": "base de chocolate, ganache de maní, bananas caramelizadas, salsa toffee y mousse de chocolate", "precio": 7500, "imagen": "torta-banofee.png"},
    ],
    "Desayunos": [
        {"nombre": "Americano", "descripcion": "Cafetería + jugo exprimido + pan de campo con palta y huevo", "precio": 14400},
        {"nombre": "Tentación", "descripcion": "Café con leche + jugo exprimido + opción de pastelería", "precio": 13800},
        {"nombre": "Criollo", "descripcion": "Cafetería + jugo exprimido + tostadas con queso, manteca o dulce", "precio": 10800},
        {"nombre": "Alfajor", "descripcion": "Cafetería + jugo exprimido + alfajor", "precio": 10700},
    ],
    "Promos": [
        {"nombre": "Clásico", "descripcion": "Infusión cafetería + 2 medialunas o tostadas + jugo", "precio": 7800},
        {"nombre": "Minis", "descripcion": "Chocolatada / Jugo mediano + cookies / sandwich", "precio": 9000},
        {"nombre": "Criollo", "descripcion": "Infusión cafetería + tostada con manteca o queso y dulce + jugo", "precio": 8300},
        {"nombre": "Alfajor", "descripcion": "Infusión cafetería + alfajor a elección + jugo", "precio": 8000},
        {"nombre": "Tentación", "descripcion": "Café especial + porción de torta + jugo", "precio": 14500},
        {"nombre": "Saludable", "descripcion": "Infusión cafetería + yogurt con granola + frutillas y arándanos + jugo", "precio": 13800},
        {"nombre": "Americano", "descripcion": "Infusión cafetería + tostada con huevo y palta + jugo", "precio": 12800},
        {"nombre": "Stella", "descripcion": "Cerveza + focaccia a elección", "precio": 16000},
        {"nombre": "Carlitos", "descripcion": "Infusión cafetería + tostada con jamón y queso + jugo", "precio": 16600},
        {"nombre": "Gran Americano", "descripcion": "Infusión cafetería + jamón, queso y huevo + mix de frutas + tostadas + jugo", "precio": 17400},
        {"nombre": "Trío Macarons", "descripcion": "Té en hebras / Infusión cafetería + jugo + trío de macarons", "precio": 12800},
        {"nombre": "Tina", "descripcion": "Tetera té en hebras + 2 porciones budín + 2 scones + madeleine + financieros + 2 jugos (para compartir)", "precio": 17400},
        {"nombre": "Degustación", "descripcion": "jarra de limonada a elección + degustación de focaccia (para compartir)", "precio": 37200},
    ],
}


def build_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="Café Tradicional")
    ws.append(HEADERS_CAFE)
    for item in CAFE_TRADICIONAL:
        ws.append([item.get(col, "si" if col == "disponible" else None) for col in HEADERS_CAFE])

    for hoja, items in SECTORES.items():
        ws = wb.create_sheet(title=hoja)
        ws.append(HEADERS_ROW)
        for item in items:
            ws.append([item.get(col, "si" if col == "disponible" else None) for col in HEADERS_ROW])

    return wb


def build_fallback():
    fallback = {"Café Tradicional": [dict(item, disponible=item.get("disponible", "si")) for item in CAFE_TRADICIONAL]}
    for hoja, items in SECTORES.items():
        fallback[hoja] = [dict(item, disponible=item.get("disponible", "si")) for item in items]
    return fallback


def main():
    wb = build_workbook()
    xlsx_path = os.path.join(BASE_DIR, "menu.xlsx")
    wb.save(xlsx_path)
    print(f"Escrito {xlsx_path}")

    fallback_path = os.path.join(BASE_DIR, "assets", "menu-fallback.json")
    with open(fallback_path, "w", encoding="utf-8") as f:
        json.dump(build_fallback(), f, ensure_ascii=False, indent=2)
    print(f"Escrito {fallback_path}")


if __name__ == "__main__":
    main()
