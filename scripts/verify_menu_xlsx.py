#!/usr/bin/env python3
"""Verifica que menu.xlsx tenga las 13 hojas esperadas, con encabezados
correctos y valores puntuales que coinciden con los precios vigentes."""

import os

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HEADERS_ROW = ["nombre", "descripcion", "disponible", "compartir", "precio", "precio_2", "imagen"]
HEADERS_CAFE = ["nombre", "disponible", "precio_s", "precio_m", "precio_l", "precio_xl"]

EXPECTED_SHEETS = [
    "Café Tradicional", "Café de Especialidad", "Materos", "Frozzen",
    "Té en Hebras", "Frapucchino", "Bebidas", "Pastelería Propia",
    "Antojos", "Focaccias Rellenas", "Pastelería Moderna", "Desayunos",
    "Promos",
]


def main():
    path = os.path.join(BASE_DIR, "menu.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)

    assert wb.sheetnames == EXPECTED_SHEETS, (
        f"Hojas inesperadas.\nEsperado: {EXPECTED_SHEETS}\nObtenido: {wb.sheetnames}"
    )

    cafe = wb["Café Tradicional"]
    header_cafe = [c.value for c in next(cafe.iter_rows(min_row=1, max_row=1))]
    assert header_cafe == HEADERS_CAFE, f"Encabezado incorrecto en 'Café Tradicional': {header_cafe}"
    fila2 = [c.value for c in next(cafe.iter_rows(min_row=2, max_row=2))]
    assert fila2[0] == "Expresso" and fila2[1] == "si", fila2
    assert fila2[2] == 3300 and fila2[3] == 4200 and fila2[4] == 4800 and fila2[5] == 5500, fila2

    for hoja in EXPECTED_SHEETS[1:]:
        ws = wb[hoja]
        header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(HEADERS_ROW)))]
        assert header_row == HEADERS_ROW, f"Encabezado incorrecto en '{hoja}': {header_row}"
        assert ws.max_row >= 2, f"La hoja '{hoja}' no tiene filas de datos"

    bebidas = wb["Bebidas"]
    filas = list(bebidas.iter_rows(min_row=2, max_col=len(HEADERS_ROW), values_only=True))
    jugo = next(f for f in filas if f[0] == "Jugo natural de naranja")
    assert jugo[4] == 6000 and jugo[5] == 7600, jugo

    moderna = wb["Pastelería Moderna"]
    fila2 = [c.value for c in next(moderna.iter_rows(min_row=2, max_row=2, max_col=len(HEADERS_ROW)))]
    assert fila2[0] == "Hawaiana", fila2
    assert fila2[6] == "torta-hawaiana.png", fila2

    promos = wb["Promos"]
    filas_promos = list(promos.iter_rows(min_row=2, max_col=len(HEADERS_ROW), values_only=True))
    filas_promos = [f for f in filas_promos if f[0]]
    nombres_promos = [f[0] for f in filas_promos]
    assert len(nombres_promos) == 13, f"Se esperaban 13 promos, hay {len(nombres_promos)}"
    assert "Agrandá a tazón (extra)" not in nombres_promos, "La constante del tazón extra no debe estar mezclada en la tabla de promos"

    comparten = {f[0] for f in filas_promos if f[3] == "si"}
    assert comparten == {"Tina", "Degustación"}, f"Se esperaba compartir=si sólo en Tina y Degustación, se encontró: {comparten}"

    header_completo = [c.value for c in next(promos.iter_rows(min_row=1, max_row=1))]
    idx_nombre_k = header_completo.index("nombre_constante")
    idx_precio_k = header_completo.index("precio_constante")
    fila_constante = [c.value for c in next(promos.iter_rows(min_row=2, max_row=2))]
    assert fila_constante[idx_nombre_k] == "Agrandá a tazón (extra)", fila_constante
    assert fila_constante[idx_precio_k] == 1500, fila_constante

    print("OK: menu.xlsx tiene las 13 hojas, encabezados y valores esperados.")


if __name__ == "__main__":
    main()
